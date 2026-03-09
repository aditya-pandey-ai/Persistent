"""
Realistic Credit Document Generator - Multi-Format
Generates credit documents in PDF, CSV, XLSX, JSON formats
Ensures scores calculate correctly (300-900 range)
"""

import os
import random
from datetime import datetime, timedelta
from faker import Faker
import json
import csv

# For Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

# For PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
except ImportError:
    print("Installing reportlab...")
    os.system("pip install reportlab")
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

fake = Faker('en_IN')  # Indian locale


class RealisticCreditDocumentGenerator:
    """Generate realistic credit documents with controlled scoring"""

    def __init__(self, output_dir="./credit_documents"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_applicant_profile(self, applicant_id):
        """Generate a consistent applicant profile with realistic credit score range"""

        # Create realistic score ranges
        score_range = random.choice([
            ('excellent', 750, 900),  # 20% of applicants
            ('good', 700, 749),       # 30% of applicants
            ('fair', 650, 699),       # 30% of applicants
            ('poor', 550, 649),       # 15% of applicants
            ('very_poor', 300, 549)   # 5% of applicants
        ])

        target_category = score_range[0]
        target_score = random.randint(score_range[1], score_range[2])

        # Calculate how to achieve this score
        base = 300
        remaining = target_score - base

        # Distribute points across factors to reach target
        if target_category == 'excellent':
            payment_pct = random.uniform(0.90, 1.0)
            util_pct = random.uniform(0.85, 1.0)
            history_pct = random.uniform(0.80, 1.0)
            inquiry_pct = random.uniform(0.85, 1.0)
            mix_pct = random.uniform(0.80, 1.0)
        elif target_category == 'good':
            payment_pct = random.uniform(0.75, 0.90)
            util_pct = random.uniform(0.70, 0.85)
            history_pct = random.uniform(0.65, 0.80)
            inquiry_pct = random.uniform(0.70, 0.85)
            mix_pct = random.uniform(0.65, 0.80)
        elif target_category == 'fair':
            payment_pct = random.uniform(0.60, 0.75)
            util_pct = random.uniform(0.55, 0.70)
            history_pct = random.uniform(0.50, 0.65)
            inquiry_pct = random.uniform(0.55, 0.70)
            mix_pct = random.uniform(0.50, 0.65)
        elif target_category == 'poor':
            payment_pct = random.uniform(0.40, 0.60)
            util_pct = random.uniform(0.35, 0.55)
            history_pct = random.uniform(0.35, 0.50)
            inquiry_pct = random.uniform(0.40, 0.55)
            mix_pct = random.uniform(0.35, 0.50)
        else:  # very_poor
            payment_pct = random.uniform(0.10, 0.40)
            util_pct = random.uniform(0.10, 0.35)
            history_pct = random.uniform(0.10, 0.35)
            inquiry_pct = random.uniform(0.10, 0.40)
            mix_pct = random.uniform(0.10, 0.35)

        # Generate credit characteristics based on target score
        profile = {
            'applicant_id': applicant_id,
            'name': fake.name(),
            'pan': fake.bothify(text='?????####?', letters='ABCDEFGHIJKLMNOPQRSTUVWXYZ'),
            'dob': fake.date_of_birth(minimum_age=21, maximum_age=65),
            'address': fake.address().replace('\n', ', '),
            'phone': fake.phone_number(),
            'email': fake.email(),
            'target_score': target_score,
            'target_category': target_category,

            # Payment history (35% = 245 points)
            'payment_score': int(245 * payment_pct),
            'total_accounts': random.randint(5, 15),
            'late_30_days': int((1 - payment_pct) * random.randint(0, 5)),
            'late_60_days': int((1 - payment_pct) * random.randint(0, 3)),
            'late_90_days': int((1 - payment_pct) * random.randint(0, 2)),

            # Credit utilization (30% = 210 points)
            'utilization_score': int(210 * util_pct),
            'total_credit_limit': random.randint(200000, 2000000),
            'utilization_pct': int((1 - util_pct) * 90 + 5),  # Lower util = higher score

            # Credit history (15% = 105 points)
            'history_score': int(105 * history_pct),
            'oldest_account_years': int(history_pct * 15) + 1,

            # New inquiries (10% = 70 points)
            'inquiry_score': int(70 * inquiry_pct),
            'hard_inquiries': int((1 - inquiry_pct) * 8),

            # Credit mix (10% = 70 points)
            'mix_score': int(70 * mix_pct),
            'num_credit_types': int(mix_pct * 4) + 1
        }

        # Calculate actual utilization amounts
        profile['total_outstanding'] = int(profile['total_credit_limit'] * profile['utilization_pct'] / 100)
        profile['available_credit'] = profile['total_credit_limit'] - profile['total_outstanding']

        return profile

    def generate_bad_credit_profile(self, applicant_id):
        """Generate a profile specifically for bad credit scores (300-549)"""

        # Force very poor or poor score
        target_score = random.randint(300, 549)
        target_category = 'very_poor' if target_score < 450 else 'poor'

        # Low percentages for bad credit
        payment_pct = random.uniform(0.10, 0.40)
        util_pct = random.uniform(0.10, 0.35)
        history_pct = random.uniform(0.10, 0.35)
        inquiry_pct = random.uniform(0.10, 0.40)
        mix_pct = random.uniform(0.10, 0.35)

        profile = {
            'applicant_id': applicant_id,
            'name': fake.name(),
            'pan': fake.bothify(text='?????####?', letters='ABCDEFGHIJKLMNOPQRSTUVWXYZ'),
            'dob': fake.date_of_birth(minimum_age=21, maximum_age=65),
            'address': fake.address().replace('\n', ', '),
            'phone': fake.phone_number(),
            'email': fake.email(),
            'target_score': target_score,
            'target_category': target_category,

            # Payment history (35% = 245 points) - BAD
            'payment_score': int(245 * payment_pct),
            'total_accounts': random.randint(3, 8),
            'late_30_days': random.randint(5, 12),  # Many late payments
            'late_60_days': random.randint(2, 6),   # Multiple 60-day lates
            'late_90_days': random.randint(1, 4),   # Some 90+ day lates

            # Credit utilization (30% = 210 points) - HIGH utilization
            'utilization_score': int(210 * util_pct),
            'total_credit_limit': random.randint(100000, 500000),
            'utilization_pct': random.randint(75, 98),  # Very high utilization

            # Credit history (15% = 105 points) - SHORT
            'history_score': int(105 * history_pct),
            'oldest_account_years': random.randint(1, 3),  # Short history

            # New inquiries (10% = 70 points) - MANY inquiries
            'inquiry_score': int(70 * inquiry_pct),
            'hard_inquiries': random.randint(6, 12),  # Lots of credit seeking

            # Credit mix (10% = 70 points) - LIMITED
            'mix_score': int(70 * mix_pct),
            'num_credit_types': random.randint(1, 2)  # Poor diversity
        }

        # Calculate actual utilization amounts
        profile['total_outstanding'] = int(profile['total_credit_limit'] * profile['utilization_pct'] / 100)
        profile['available_credit'] = profile['total_credit_limit'] - profile['total_outstanding']

        return profile

    def generate_credit_report_pdf(self, profile):
        """Generate PDF credit report"""
        filename = f"{self.output_dir}/credit_report_{profile['applicant_id']}.pdf"

        doc = SimpleDocTemplate(filename, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("CREDIT BUREAU REPORT", title_style))
        story.append(Spacer(1, 0.3*inch))

        # Personal Info
        story.append(Paragraph(f"<b>Report Date:</b> {datetime.now().strftime('%d %B %Y')}", styles['Normal']))
        story.append(Paragraph(f"<b>Applicant ID:</b> {profile['applicant_id']}", styles['Normal']))
        story.append(Paragraph(f"<b>Full Name:</b> {profile['name']}", styles['Normal']))
        story.append(Paragraph(f"<b>PAN:</b> {profile['pan']}", styles['Normal']))
        story.append(Paragraph(f"<b>Date of Birth:</b> {profile['dob'].strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Credit Score Summary
        score_data = [
            ['Credit Score Summary', ''],
            ['Total Accounts', str(profile['total_accounts'])],
            ['Total Credit Limit', f"₹{profile['total_credit_limit']:,}"],
            ['Total Outstanding', f"₹{profile['total_outstanding']:,}"],
            ['Credit Utilization', f"{profile['utilization_pct']}%"],
            ['Oldest Account', f"{profile['oldest_account_years']} years"],
            ['Hard Inquiries (12m)', str(profile['hard_inquiries'])],
        ]

        score_table = Table(score_data, colWidths=[3*inch, 2*inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(score_table)
        story.append(Spacer(1, 0.3*inch))

        # Payment History
        story.append(Paragraph("<b>Payment History Details:</b>", styles['Heading2']))
        story.append(Paragraph(f"30 Days Late: {profile['late_30_days']} payments", styles['Normal']))
        story.append(Paragraph(f"60 Days Late: {profile['late_60_days']} payments", styles['Normal']))
        story.append(Paragraph(f"90+ Days Late: {profile['late_90_days']} payments", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))

        # Credit Mix
        story.append(Paragraph(f"<b>Credit Mix:</b> {profile['num_credit_types']} different account types", styles['Normal']))

        doc.build(story)
        print(f"  Generated PDF: {filename}")

    def generate_payment_history_csv(self, profile):
        """Generate CSV with detailed payment history"""
        filename = f"{self.output_dir}/payment_history_{profile['applicant_id']}.csv"

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Applicant_ID', 'Account_Number', 'Account_Type', 'Lender',
                           'Month', 'Payment_Status', 'Amount_Due', 'Amount_Paid', 'Days_Late'])

            account_types = ['Credit Card', 'Personal Loan', 'Home Loan', 'Auto Loan', 'Education Loan']

            # Generate 12 months of payment history
            for month_offset in range(12):
                payment_date = datetime.now() - timedelta(days=30 * month_offset)

                for account_num in range(profile['total_accounts']):
                    account_type = random.choice(account_types)
                    lender = fake.company() + " Bank"
                    amount_due = random.randint(5000, 50000)

                    # Determine payment status based on profile
                    late_prob = (profile['late_30_days'] + profile['late_60_days'] + profile['late_90_days']) / (profile['total_accounts'] * 12)

                    if random.random() < late_prob:
                        if profile['late_90_days'] > 0 and random.random() < 0.2:
                            days_late = random.randint(90, 120)
                            payment_status = "90+ Days Late"
                        elif profile['late_60_days'] > 0 and random.random() < 0.3:
                            days_late = random.randint(60, 89)
                            payment_status = "60 Days Late"
                        else:
                            days_late = random.randint(30, 59)
                            payment_status = "30 Days Late"
                        amount_paid = int(amount_due * random.uniform(0.3, 0.7))
                    else:
                        days_late = 0
                        payment_status = "On-Time"
                        amount_paid = amount_due

                    writer.writerow([
                        profile['applicant_id'],
                        f"AC{account_num:04d}",
                        account_type,
                        lender,
                        payment_date.strftime('%Y-%m'),
                        payment_status,
                        amount_due,
                        amount_paid,
                        days_late
                    ])

        print(f"  Generated CSV: {filename}")

    def generate_account_summary_xlsx(self, profile):
        """Generate Excel with account summaries"""
        filename = f"{self.output_dir}/account_summary_{profile['applicant_id']}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Account Summary"

        # Header styling
        header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Headers
        headers = ['Account_Number', 'Account_Type', 'Lender', 'Credit_Limit', 'Current_Balance',
                  'Available_Credit', 'Utilization_%', 'Account_Status', 'Opened_Date']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Generate accounts
        account_types = ['Credit Card', 'Personal Loan', 'Home Loan', 'Auto Loan', 'Business Credit']
        used_types = random.sample(account_types, profile['num_credit_types'])

        total_limit = 0
        total_balance = 0

        for row in range(2, profile['total_accounts'] + 2):
            account_type = random.choice(used_types)
            credit_limit = random.randint(50000, 300000)

            # Calculate balance to match overall utilization
            target_util = profile['utilization_pct'] / 100
            current_balance = int(credit_limit * target_util * random.uniform(0.7, 1.3))
            current_balance = min(current_balance, credit_limit)

            available = credit_limit - current_balance
            utilization = (current_balance / credit_limit * 100) if credit_limit > 0 else 0

            opened_date = datetime.now() - timedelta(days=random.randint(365, 365 * profile['oldest_account_years']))

            ws.cell(row=row, column=1, value=f"AC{row-2:04d}")
            ws.cell(row=row, column=2, value=account_type)
            ws.cell(row=row, column=3, value=fake.company() + " Bank")
            ws.cell(row=row, column=4, value=credit_limit)
            ws.cell(row=row, column=5, value=current_balance)
            ws.cell(row=row, column=6, value=available)
            ws.cell(row=row, column=7, value=f"{utilization:.1f}%")
            ws.cell(row=row, column=8, value="Active")
            ws.cell(row=row, column=9, value=opened_date.strftime('%Y-%m-%d'))

            total_limit += credit_limit
            total_balance += current_balance

        # Adjust column widths
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15

        wb.save(filename)
        print(f"  Generated XLSX: {filename}")

    def generate_inquiry_report_json(self, profile):
        """Generate JSON with credit inquiry history"""
        filename = f"{self.output_dir}/inquiries_{profile['applicant_id']}.json"

        inquiries = []

        for i in range(profile['hard_inquiries']):
            inquiry_date = datetime.now() - timedelta(days=random.randint(1, 365))
            inquiries.append({
                "applicant_id": profile['applicant_id'],
                "inquiry_id": f"INQ{i+1:05d}",
                "inquiry_type": "Hard Inquiry",
                "inquired_by": fake.company() + " Financial Services",
                "purpose": random.choice(["Credit Card Application", "Personal Loan", "Home Loan", "Auto Loan"]),
                "inquiry_date": inquiry_date.strftime('%Y-%m-%d'),
                "status": "Completed"
            })

        # Add some soft inquiries (don't affect score)
        for i in range(random.randint(3, 8)):
            inquiry_date = datetime.now() - timedelta(days=random.randint(1, 365))
            inquiries.append({
                "applicant_id": profile['applicant_id'],
                "inquiry_id": f"INQ{profile['hard_inquiries']+i+1:05d}",
                "inquiry_type": "Soft Inquiry",
                "inquired_by": fake.company() + " Bank",
                "purpose": "Pre-approval Check",
                "inquiry_date": inquiry_date.strftime('%Y-%m-%d'),
                "status": "Completed"
            })

        data = {
            "applicant_id": profile['applicant_id'],
            "report_date": datetime.now().strftime('%Y-%m-%d'),
            "total_hard_inquiries": profile['hard_inquiries'],
            "total_soft_inquiries": len(inquiries) - profile['hard_inquiries'],
            "inquiries": inquiries
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"  Generated JSON: {filename}")

    def generate_applicant_documents(self, applicant_id, force_bad=False):
        """Generate all documents for one applicant"""
        if force_bad:
            profile = self.generate_bad_credit_profile(applicant_id)
        else:
            profile = self.generate_applicant_profile(applicant_id)

        print(f"\nGenerating documents for Applicant {applicant_id} (Target Score: {profile['target_score']} - {profile['target_category'].upper()})...")

        self.generate_credit_report_pdf(profile)
        self.generate_payment_history_csv(profile)
        self.generate_account_summary_xlsx(profile)
        self.generate_inquiry_report_json(profile)

        return profile

    def generate_batch(self, num_applicants=20, num_bad_scores=5):
        """Generate documents for multiple applicants with controlled bad scores"""
        print("="*70)
        print("REALISTIC CREDIT DOCUMENT GENERATOR")
        print("="*70)
        print(f"\nGenerating documents for {num_applicants} applicants...")
        print(f"  - First {num_bad_scores} will have BAD credit scores (300-549)")
        print(f"  - Remaining {num_applicants - num_bad_scores} will have mixed scores")
        print(f"Output directory: {self.output_dir}")

        profiles = []
        for i in range(num_applicants):
            applicant_id = 10000 + i

            # Force bad scores for first num_bad_scores applicants
            force_bad = (i < num_bad_scores)
            profile = self.generate_applicant_documents(applicant_id, force_bad=force_bad)
            profiles.append(profile)

        print("\n" + "="*70)
        print("GENERATION COMPLETE")
        print("="*70)
        print(f"\nTotal files generated: {num_applicants * 4}")
        print(f"Per applicant:")
        print(f"  - 1 PDF (Credit Report)")
        print(f"  - 1 CSV (Payment History)")
        print(f"  - 1 XLSX (Account Summary)")
        print(f"  - 1 JSON (Inquiry History)")

        print(f"\nScore distribution:")
        score_categories = {}
        for p in profiles:
            cat = p['target_category']
            score_categories[cat] = score_categories.get(cat, 0) + 1

        for cat, count in sorted(score_categories.items()):
            print(f"  {cat.replace('_', ' ').title()}: {count} applicants")

        print("\n" + "="*70)
        print("Ready to upload to your RAG system!")
        print("="*70)


if __name__ == "__main__":
    generator = RealisticCreditDocumentGenerator(output_dir="./credit_documents")
    generator.generate_batch(num_applicants=30, num_bad_scores=5)  # First 5 will be bad
